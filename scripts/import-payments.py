import sqlite3
import openpyxl
import os
import datetime
import re
from datetime import datetime as dt

DB_PATH = '/home/diegopalhano/.openclaw/workspace/data/atlas-memory.db'
EXCEL_PATH = '/home/diegopalhano/Downloads/Payments.xlsx'

HOUSE_MAP = {
    '43 Redfern St': 'WL4',
    '36 Rosa St': 'SH2',
    '28 Taylor St': 'WL3',
    '15 Cameron St': 'SB1',
    '553 Vulture St': 'EB1',
    '3 Hardgrave St': 'WE1',
    '40 Rosa St': 'SH1',
    '37 Marian St': 'CO1',
    '111 Juliette St': 'GS1',
    '606 Vulture St': 'EB2',
    '69 Gresham': 'EB3',
    '147 Warry St': 'V5',
    '41 Park St': 'SH3',
    '79 Albert St': 'BRIS1',
    '50 Peninsular': 'SP9'
}

SKIP_SHEETS = ['Control', 'Bank', 'Sheet1', 'Contract']

def slugify(text):
    return re.sub(r'[^a-z0-9]', '_', text.lower())

def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS payments (
          id TEXT PRIMARY KEY,
          tenant_name TEXT NOT NULL,
          house_code TEXT NOT NULL,
          property_address TEXT NOT NULL,
          room_number INTEGER,
          weekly_rent REAL,
          week_due TEXT NOT NULL,
          paid_date TEXT,
          status TEXT NOT NULL,
          notes TEXT,
          imported_at TEXT NOT NULL
        )
    ''')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_payments_tenant ON payments(tenant_name)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_payments_house ON payments(house_code)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_payments_week ON payments(week_due)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_payments_status ON payments(status)')
    conn.commit()
    return conn

def import_payments():
    conn = init_db()
    cursor = conn.cursor()
    
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: {EXCEL_PATH} not found.")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    total_imported = 0
    properties_processed = 0
    now_iso = dt.now().isoformat()

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        
        sheet = wb[sheet_name]
        property_address = sheet['A1'].value
        if not property_address:
            # Fallback to sheet name if A1 is empty
            property_address = sheet_name
        
        # Strip potential extra info from A1 (sometimes it has "Property: Name")
        property_address = str(property_address).replace('Property:', '').strip()
        
        house_code = HOUSE_MAP.get(property_address)
        if not house_code:
            # Try fuzzy match or skip
            # print(f"Skipping unknown property: {property_address}")
            continue

        properties_processed += 1
        
        # Row 1 has dates from Column D onwards
        date_headers = {}
        for col in range(4, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            if isinstance(val, (dt, datetime.date)):
                date_headers[col] = val.strftime('%Y-%m-%d')
            elif isinstance(val, str):
                # Try to parse date string if it's not a datetime object
                try:
                    # Adjust format as needed based on actual Excel strings
                    d = dt.strptime(val, '%Y-%m-%d')
                    date_headers[col] = d.strftime('%Y-%m-%d')
                except:
                    pass

        # Row 3 onwards: Tenants
        for row in range(3, sheet.max_row + 1):
            tenant_name = sheet.cell(row=row, column=1).value
            if not tenant_name:
                continue
            
            room_number = sheet.cell(row=row, column=2).value
            weekly_rent = sheet.cell(row=row, column=3).value
            
            for col, week_due in date_headers.items():
                cell_val = sheet.cell(row=row, column=col).value
                
                paid_date = None
                status = 'missing'
                notes = None
                
                if isinstance(cell_val, (dt, datetime.date)):
                    status = 'paid'
                    paid_date = cell_val.strftime('%Y-%m-%d')
                elif cell_val == 'EMPTY':
                    status = 'empty'
                elif cell_val == 'LEFT':
                    status = 'left'
                elif cell_val == 'BOND':
                    status = 'bond'
                elif cell_val == 'missing':
                    status = 'missing'
                elif cell_val is None:
                    # Check if week_due is in the past
                    if week_due < dt.now().strftime('%Y-%m-%d'):
                        status = 'missing'
                    else:
                        # Future week, skip
                        continue
                else:
                    # Treat anything else as notes or paid if it looks like a date
                    notes = str(cell_val)
                
                # Generate unique ID
                slug_tenant = slugify(str(tenant_name))
                record_id = f"{house_code}_{slug_tenant}_{week_due}"[:80]
                
                cursor.execute('''
                    INSERT INTO payments (
                        id, tenant_name, house_code, property_address, 
                        room_number, weekly_rent, week_due, paid_date, 
                        status, notes, imported_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(id) DO UPDATE SET
                        paid_date=excluded.paid_date,
                        status=excluded.status,
                        notes=excluded.notes,
                        imported_at=excluded.imported_at
                ''', (
                    record_id, str(tenant_name), house_code, property_address,
                    room_number, weekly_rent, week_due, paid_date,
                    status, notes, now_iso
                ))
                total_imported += 1

    conn.commit()
    conn.close()
    print(f"Imported {total_imported} records across {properties_processed} properties.")

if __name__ == '__main__':
    import_payments()
